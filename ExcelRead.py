import openpyxl as O


class Excel:
    def __init__(self):
        self.excel_file = "C:\data.xlsx"
        self.excel_worksheet = 'Sheet1'
        self.wb = O.load_workbook(self.excel_file)
        self.ws = self.wb[self.excel_worksheet]

    def value_finder(self, row, column):
        return self.ws.cell(row, column).value

    def passed_test(self, row, column):
        self.ws.cell(row, column).value = 'passed'
        self.wb.save(self.excel_file)

    def failed_test(self, row, column):
        self.ws.cell(row, column).value = 'failed'
        self.wb.save(self.excel_file)

    def delete_cell(self, row, column):
        self.ws.cell(row, column).value = None
        self.wb.save(self.excel_file)


e = Excel()
# # print(e.value_finder(26, 3))
# e.delete_cell(26, 12)
print(e.value_finder(26, 3))
