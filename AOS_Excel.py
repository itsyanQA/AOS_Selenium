import openpyxl as O
from openpyxl.styles import PatternFill


class Excel:
    def __init__(self):
        '''we assign our excel file to a variable, as well as the worksheet.'''
        self.excel_file = "C:\data.xlsx"
        self.excel_worksheet = 'Sheet1'
        self.wb = O.load_workbook(self.excel_file)
        self.ws = self.wb[self.excel_worksheet]

    def value_finder(self, row, column):
        '''Gets row and a column, returns the value of the specific cell'''
        return self.ws.cell(row, column).value

    def passed_test(self, row, column):
        '''Gets row and column, writes "passed" and puts the color green on every located cell'''
        self.ws.cell(row, column).value = 'passed'
        green = PatternFill(patternType='solid', fgColor='35FC03')
        self.ws.cell(row, column).fill = green
        self.wb.save(self.excel_file)

    def delete_cell(self, row, column):
        '''Gets row and column, assigns none to the cell'''
        self.ws.cell(row, column).value = None
        self.wb.save(self.excel_file)

    def red_fill(self, location):
        '''Fills the located cell with color red'''
        red = PatternFill(patternType='solid', fgColor='FC2C03')
        self.ws[location].fill = red
        self.wb.save(self.excel_file)

    def pass_fail_status(self):
        '''Prints the results of all the tests'''
        range = self.ws['C26':'M26']
        for cell in range:
            for value in cell:
                print(f"{value}: {value.value}")

    def fail_everything(self):
        '''We assign a range and for every value in the cell we put the word "Failed".'''
        range = self.ws['C26':'M26']
        for cell in range:
            for value in cell:
                value.value = 'failed'
                self.wb.save(self.excel_file)









